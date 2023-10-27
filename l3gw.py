import openpyxl as xl


wb1 = xl.load_workbook('L3 Glue Weights.xlsx')
read = wb1['Adhesive Weight']
wb2 = xl.load_workbook('L3 GW New.xlsx')
write = wb2.active

numrows = 110
numcols = 50
writecol = 3
i = 3
for r in range(6,numrows):
    date = read.cell(row=r, column=1).value
    for c in range(1,numcols):
        print(c)
        if c == 1:
            write.cell(row=write.max_row+1, column=1).value = date
            write.cell(row=write.max_row, column=2).value = write.cell(row=i, column=7).value
            i += 1
            continue
        if (c-2)%2 == 0 and c != 2:
            write.cell(row=write.max_row+1, column=1).value = date
            write.cell(row=write.max_row, column=2).value = write.cell(row=i, column=7).value
            i += 1
        try:
            write.cell(row=write.max_row, column=writecol).value = float(read.cell(row=r, column=c).value)
        except:
            print("Not a number")
        writecol += 1
        if writecol == 5:
            writecol = 3
        if i == 27:
            i = 3
wb2.save('L3 GW New.xlsx')
