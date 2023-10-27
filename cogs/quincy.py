import os
import PIL
from PIL import Image, ImageEnhance, ImageDraw, ImageFont
import zipfile
import pytesseract
from pdf2image import convert_from_path, convert_from_bytes
import openpyxl as xl
from openpyxl.styles import Alignment
import datetime

pytesseract.pytesseract.tesseract_cmd = r'C:\Users\asuresh\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'

wb = openpyxl.load_workbook('test.xlsx')
sheet = wb.active

file = 'quincy.zip'
zipfile_data=zipfile.ZipFile(file)
zipfile_data.extractall()

ratio = 2200/1700
width = 2350
height = width*ratio

for x in range(len(zipfile_data.namelist())):
    convertedimage = convert_from_path(zipfile_data.namelist()[x])
    resizedim = convertedimage[0].resize((int(width),int(height)))
    imtext = pytesseract.image_to_string(resizedim)
    items = imtext.split()

    print('Entering data in for file: ' + zipfile_data.namelist()[x])

    specs = {}
    specs['RC#'] = items[items.index('Vehicle')+2] + items[items.index('Vehicle')+3]
    specs['lot#'] = 'QCY' + items[items.index('Lot')+3]
    specs['ship date'] = itms[items.index('Ship')+2]
    specs['insol'] = items[items.index('Acid')+4]
    specs['a'] = items[items.index('a*')+3]
    specs['b'] = items[items.index('b*')+3]
    specs['L'] = items[items.index('L*')+3]
    specs['200mesh'] = items[items.index('200')+4]
    specs['40mesh'] = items[items.index('40')+4]
    specs['moist'] = items[items.index('Moisture(%)')+5]
    counter = 0
    for i in specs:
        try:
            float(specs[i])
        except:
            counter += 1
    if counter >= 7:
        print('Words pulled from pdf were unable to be parsed')
        print('File must be entered in manually')
        print('\n')
        continue
    y = 1
    first_row = 0
    delete = 1
    for i in specs:
        print(specs[i],end = ' ')
        if first_row == 0:
            sheet.cell(row=sheet.max_row+1,column=y).value = specs[i]
            first_row = 1
        else:
            if y>3:
                try:
                    sheet.cell(row=sheet.max_row,column=y).value = float(specs[i])
                except ValueError:
                    delete = 0
                    print("\nCould not convert " + specs[i] + " to a float")
                    print('Data will be entered as text')
                    sheet.cell(row=sheet.max_row,column=y).value = specs[i]
            else:
                sheet.cell(row=sheet.max_row,column=y).value = specs[i]
        y = y + 1
    print('\n')
    if os.path.exists(zipfile_data.namelist()[x]) and delete == 1:
        os.remove(zipfile_data.namelist()[x])
wb.save('test.xlsx')
print('Files that presented issues were not deleted and should be checked/inputted manually')
