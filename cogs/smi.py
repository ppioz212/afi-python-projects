import os
import PIL
from PIL import Image, ImageEnhance, ImageDraw, ImageFont
import zipfile
import pytesseract
from pdf2image import convert_from_path, convert_from_bytes
import openpyxl as xl
from openpyxl.styles import Alignment
import datetime
pytesseract.pytesseract.tesseract_cmd = r'C:\Users\asuresh\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'

wb = xl.load_workbook('LIMESTONE - SMI.xlsx')
sheet = wb['SMI DATA']

file = 'smi.zip'
zipfile_data=zipfile.ZipFile(file)
zipfile_data.extractall()

ratio = 2200/1700
width = 2350
height = width*ratio

sl = 139
index_values = [sl-100,sl-99,sl-93,sl-92,sl-72,sl-64,sl-56,sl-50,sl-44,sl-38,sl-26,sl-20,sl-14,sl-10]
index_names = ['ship date','bol#','lot#','RC#','drybr','a','b','16m','50m','100m','-200m','insol','%mois','td']
indices = list(zip(index_names,index_values))
def errormessage(x):
    print('Words pulled from ' + str(x) + ' were unable to be parsed')
    print('File must be entered in manually')
    print('\n')

for pdf_filename in zipfile_data.namelist():
    convertedimage = convert_from_path(pdf_filename)
    resizedim = convertedimage[0].resize((int(width),int(height)))
    imtext = pytesseract.image_to_string(resizedim)
    items = imtext.split()
    if len(items) == sl:
        try:
            extra = items[index_values[2]].split(':')
            items.remove(items[index_values[2]])
            items.insert(index_values[2],extra[0])
            items.insert(index_values[2]+1,extra[1])
        except:
            errormessage(pdf_filename)
            continue
    elif len(items) == sl + 1:
        if items[index_values[2]][len(items[index_values[2]])-1] == ':':
            extra = items[index_values[2]].split(':')
            items.remove(items[index_values[2]])
            items.insert(index_values[2],extra[0])
        elif items[index_values[2]+1][0] == ':':
            extra = items[index_values[2]+1].split(':')
            items.remove(items[index_values[2]+1])
            items.insert(index_values[2]+1,extra[1])
        else:
            errormessage(pdf_filename)
            continue
    elif len(items) == sl + 2:
        if items[index_values[2]+1] == ':':
            items.remove(items[index_values[2]+1])
    else:
        errormessage(pdf_filename)
        continue

    specs = {}
    count = 1
    for spec, index in indices:
        specs[spec] = items[index]
        if count == 1:
            try:
                datetime_obj = datetime.datetime.strptime(specs[spec], '%d-%b-%y')
                specs[spec] = datetime_obj.date()
            except:
                pass
        if count == 13:
            if specs[spec][0]!='.':
                specs[spec] = '.' + specs[spec]
        count += 1
    counter = 0

    for i in specs:
        try:
            float(specs[i])
        except ValueError:
            counter += 1
    if counter >= 7:
        errormessage(pdf_filename)
        continue

    y = 1
    delete = True
    print('Entering data in for file: ' + pdf_filename)
    for i in specs:
        if y==13:

        print(specs[i],end = ' ')
        if y==1:
                sheet.cell(row=sheet.max_row+1,column=y).value = specs[i]
                sheet.cell(row=sheet.max_row,column=y).alignment = Alignment(horizontal='center')
            y += 1
            continue
        else:
            if y==2 or y>4:
                try:
                    sheet.cell(row=sheet.max_row,column=y).value = float(specs[i])
                    sheet.cell(row=sheet.max_row,column=y).alignment = Alignment(horizontal='center')
                except ValueError:
                    print("\nCould not convert " + specs[i] + " to a float")
                    print('Data will be entered as text')
                    sheet.cell(row=sheet.max_row,column=y).value = specs[i]
                    sheet.cell(row=sheet.max_row,column=y).alignment = Alignment(horizontal='center')
                    delete = False
            else:
                sheet.cell(row=sheet.max_row,column=y).value = specs[i]
                sheet.cell(row=sheet.max_row,column=y).alignment = Alignment(horizontal='center')
        y = y + 1
    print('\n')
    if os.path.exists(pdf_filename):
        os.remove(pdf_filename)
wb.save('LIMESTONE - SMI.xlsx')
print('Files that presented issues were not deleted and should be inputted manually')
