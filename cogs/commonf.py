import os
# import PIL
# from PIL import Image, ImageEnhance, ImageDraw, ImageFont
import zipfile
# import pytesseract
# from pdf2image import convert_from_path
import openpyxl as xl
from openpyxl.styles import Alignment
import datetime
import pdfplumber
import re
# pytesseract.pytesseract.tesseract_cmd = r'C:\Users\asuresh\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'

ratio = 2200/1700
width = 1600
height = width*ratio

def final_call(zipfile):
    if zipfile.lower() == ('smi.zip'):
        smi(zipfile)
    # if zipfile.lower() == ('quincy.zip'):
    #     quincy(zipfile)
    if zipfile.lower() == ('marblehill.zip'):
        marblehill(zipfile)
    if zipfile.lower() == ('sudarshan.zip'):
        sudarshan(zipfile)

def zipextract(zipname):
    zipdata = zipfile.ZipFile(zipname)
    zipdata.extractall()
    return zipdata

def extract_from_image(pdf_filename):
    convertedimage = convert_from_path(pdf_filename)
    print(len(convertedimage))
    return convertedimage

def pdf_text_extract(pdf_filename):
    with pdfplumber.open(pdf_filename) as pdf:
        if len(pdf.pages)<2:
            list_range = [0]
        else:
            list_range = list(range(0,len(pdf.pages)-1))
        items_list = [pdf.pages[x].extract_text().split() for x in list_range]
    return items_list

def convert_date(raw_date):
    datetime_obj = datetime.datetime.strptime(raw_date, '%d-%b-%y')
    return datetime_obj.date()
    # Old poor format. Might use again if fixed but probably not
    # date = datetime_obj.date()
    # return date.strftime('%m/%d/%Y')

def number_pattern(number_matches,items,start):
    counter = 0
    for i in range(start,len(items)-1):
        if re.search("^(-?)(0?|([1-9][0-9]*))(\\.[0-9]+)?$",items[i]) != None:
            counter += 1
        else:
            counter = 0
            continue
        if counter == number_matches:
            return i

def errormessage(pdf_filename):
    print('Words pulled from ' + str(pdf_filename) + ' were unable to be parsed')
    print('File must be entered in manually\n\n')

def countcheck(specs):
    counter = 0
    for i in specs:
        try:
            float(specs[i])
        except:
            counter += 1
    return counter >= 7

def excel_entry(specs, file_name, sheet, format):
    y = 1
    print('Entering data in for file: ' + str(file_name))
    for i in specs:
        if format == 'smi':
            nonfloat_list = [1,3,4]
        elif format == 'quincy':
            nonfloat_list = [1,2,3]
        elif format == 'marblehill':
            nonfloat_list = [1,2,3]
        elif format == '2730K':
            nonfloat_list = [1,2,3,4]
        print(specs[i],end = ' ')
        if y==1:
                sheet.cell(row=sheet.max_row+1,column=y).value = specs[i]
                sheet.cell(row=sheet.max_row,column=y).alignment = Alignment(horizontal='center')
                y += 1
                continue
        else:
            if not y in nonfloat_list:
                try:
                    sheet.cell(row=sheet.max_row,column=y).value = float(specs[i])
                    sheet.cell(row=sheet.max_row,column=y).alignment = Alignment(horizontal='center')
                except ValueError:
                    print("\nCould not convert " + specs[i] + " to a float")
                    print('Data will be entered as text')
                    sheet.cell(row=sheet.max_row,column=y).value = specs[i]
                    sheet.cell(row=sheet.max_row,column=y).alignment = Alignment(horizontal='center')
            else:
                sheet.cell(row=sheet.max_row,column=y).value = specs[i]
                sheet.cell(row=sheet.max_row,column=y).alignment = Alignment(horizontal='center')
        y = y + 1
    print('\n')
def sudarshan(zipfile):
    pass
def smi(zipfile):
    wb = xl.load_workbook('LIMESTONE - SMI.xlsx')
    sheet = wb['SMI DATA']
    zipfile_data = (zipextract(zipfile))
    print('Running data entry for SMI Limestone')

    for pdf_filename in zipfile_data.namelist():
        for items in pdf_text_extract(pdf_filename):
            specs = {}
            specs['date'] = convert_date(items[items.index('Ship')+2])
            specs['bol #'] = items[items.index('BOL')+2]
            specs['lot #'] = items[items.index('Lot:')+1].split(':')[0]
            specs['RC #'] = items[items.index('Lot:')+1].split(':')[1]
            specs['DryBr'] = items[number_pattern(3,items,items.index('BRIGHTNESS,'))-2]
            specs['A'] = items[number_pattern(3,items,items.index('A'))-2]
            specs['B'] = items[number_pattern(3,items,items.index('B'))-2]
            specs['16M'] = items[number_pattern(3,items,items.index('+16'))-2]
            specs['50M'] = items[number_pattern(3,items,items.index('+50'))-2]
            specs['100M'] = items[number_pattern(3,items,items.index('+100'))-2]
            specs['200M'] = items[number_pattern(3,items,items.index('-200'))-2]
            specs['Insol'] = items[number_pattern(3,items,items.index('ACID'))-2]
            specs['%Moist'] = items[number_pattern(1,items,items.index('MOISTURE'))]
            specs['TD'] = items[number_pattern(1,items,items.index('TAP'))]
            if countcheck(specs):
                errormessage(pdf_filename)
                print("Error at count check")
                continue
            excel_entry(specs,pdf_filename,sheet,'smi')
            if os.path.exists(pdf_filename):
                os.remove(pdf_filename)
    wb.save('LIMESTONE - SMI.xlsx')
    print('Files that presented issues were not deleted and should be inputted manually')

def quincy(zipfile):
    # wb = xl.load_workbook('LIMESTONE - HUBER QUINCY.xlsx')
    # sheet = wb['Huber Quincy']
    zipfile_data = (zipextract(zipfile))
    print('Running data entry for Quincy Limestone')

    for pdf_filename in zipfile_data.namelist():
        images = extract_from_image(pdf_filename)
        for image in images:
            resizedim = image.resize((int(width),int(height)))
            imtext = pytesseract.image_to_string(resizedim)
            items = imtext.split()

            specs = {}
            specs_index = {}
            specs['RC#'] = items[items.index('Vehicle')+2] + items[items.index('Vehicle')+3]
            specs['lot#'] = 'QCY' + items[items.index('Lot')+3]
            # specs['ship date'] = convert_date(items[items.index('Ship')+2])
            specs['insol'] = items[items.index('Acid')+4]
            specs['a'] = items[items.index('a*')+3]
            specs['b'] = items[items.index('b*')+3]
            specs['L'] = items[items.index('L*')+3]
            specs['200mesh'] = items[items.index('200')+4]
            specs['40mesh'] = items[items.index('40')+4]
            specs['moist'] = items[items.index('Moisture(%)')+5]
            specs_index['RC#'] = items.index('Vehicle')+2
            specs_index['lot#'] = items.index('Lot')+3
            # specs['ship date'] = convert_date(items[items.index('Ship')+2])
            specs_index['insol'] = items.index('Acid')+4
            specs_index['a'] = items.index('a*')+3
            specs_index['b'] = items.index('b*')+3
            specs_index['L'] = items.index('L*')+3
            specs_index['200mesh'] = items.index('200')+4
            specs_index['40mesh'] = items.index('40')+4
            specs_index['moist'] = items.index('Moisture(%)')+5
            print(len(items))
            for i in specs:
                print(specs[i] + " " + str(specs_index[i]))
            # if countcheck(specs):
            #     errormessage(pdf_filename)
            #     print("Error at count check")
            #     continue
            # excel_entry(specs, pdf_filename, sheet, 'quincy')
            if os.path.exists(pdf_filename):
                os.remove(pdf_filename)
        # wb.save('LIMESTONE - HUBER QUINCY.xlsx')
    # print('Files that presented issues were not deleted and should be inputted manually')

def marblehill(zipfile):
    wb = xl.load_workbook('LIMESTONE - HUBER MARBLE HILL.xlsx')
    sheet = wb['MarbleHill 2014-2021']
    zipfile_data = (zipextract(zipfile))
    print('Running data entry for Marble Hill Limestone')

    for pdf_filename in zipfile_data.namelist():
        for items in pdf_text_extract(pdf_filename):
            specs = {}
            specs['RC#'] = items[items.index('Vehicle')+2] + items[items.index('Vehicle')+3]
            specs['lot#'] = items[items.index('Lot')+2]
            specs['ship date'] = convert_date(items[items.index('Ship')+2])
            specs['insol'] = items[items.index('Acid')+3]
            specs['a'] = items[items.index('a*')+2]
            specs['b'] = items[items.index('b*')+2]
            specs['L'] = items[items.index('L*')+2]
            specs['100 mesh'] = items[items.index('100')+3]
            specs['200 mesh'] = items[items.index('200')+3]
            specs['50 mesh'] = items[items.index('50')+3]
            specs['retained 50 mesh'] = 100 - float(specs['50 mesh'])
            if items[items.index('Moisture(%)')+5] == "Passed":
                specs['moist'] = items[items.index('Moisture(%)')+4]
            else:
                specs['moist'] = items[items.index('Moisture(%)')+5]
            try:
                specs['Y'] = items[items.index('Y')+3]
            except:
                specs['Y'] = "N/A"
            print(specs)
            if countcheck(specs):
                errormessage(pdf_filename)
                print("Error at count check")
                continue
            excel_entry(specs, pdf_filename, sheet, 'marblehill')
            if os.path.exists(pdf_filename):
                os.remove(pdf_filename)
    wb.save('LIMESTONE - HUBER MARBLE HILL.xlsx')
    print('Files that presented issues were not deleted and should be checked/inputted manually')
