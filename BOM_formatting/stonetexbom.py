import openpyxl as xl

wb = xl.load_workbook('L5 Stonetex BOMs Test.xlsx')
r = 2
for i in wb.sheetnames:
    wb['List'].cell( row = r, column = 1).value = i
    print(i, wb[i].cell( row = 17, column = 12).value)
    r +=1
wb.save('L5 Stonetex BOMs Test.xlsx')
