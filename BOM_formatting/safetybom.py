import openpyxl as xl

wb = xl.load_workbook('L5 Safety Zone BOMs 2022.xlsx')
c = 1
for i in wb.sheetnames:
    if c<3:
        c += 1
        continue
    # if i.split('_')[-1] == 'CBO':
    #     continue
    # wb[i].cell(row = 17, column = 13).value = '=$Q$14*Standard!$E$2'
    # wb[i].cell(row = 18, column = 13).value = '=0.001767*Standard!$C$2*M17'
    # wb[i].cell(row = 19, column = 13).value = '=M17/Standard!$D$2'
    print(wb[i].cell(row = 17, column = 13).value)
# wb.save('L5 Safety Zone BOMs 2022.xlsx')
