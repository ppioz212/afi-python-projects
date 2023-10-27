import openpyxl as xl

wb = xl.load_workbook('L5 Imperial_Multi_Metric BOMs 2022.xlsx')
sheet = wb['51801 K3 L5']
c = 1
for i in wb.sheetnames:

#     for x in range(13,19):
#         for y in range(10,13):
#             try:
#                 if wb[i].cell( row = x, column = y).value[0:2] == 'BP':
#                     bp = wb[i].cell( row = x, column = y).value
#                     wb[i].cell( row = 12, column = 16).value = bp
#             except:
#                 continue
    if c<4:
        c +=1
        continue
    # wb['List'].cell( row = r, column = 1).value = i

    wb[i].cell(row = 20, column = 13).value = '=$Q$13*Standard!$E$2'
    wb[i].cell(row = 21, column = 13).value = '=0.001767*Standard!$C$2*M20'
    wb[i].cell(row = 22, column = 13).value = '=M20/Standard!$D$2'
    # r +=1
wb.save('L5 Imperial_Multi_Metric BOMs 2022.xlsx')
