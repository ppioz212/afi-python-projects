import openpyxl as xl
import os
wb_pigs = xl.load_workbook('Pigment IDs.xlsx')
pigments = [str(wb_pigs.active.cell(row=x,column=1).value) for x in range(1,wb_pigs.active.max_row + 1)]
whitepigids = ['145048']
whitepignames = ['White Pigment - Venator TR-48', 'White Pigment']
print(pigments, len(pigments))
mydir = 'C:/Users/asuresh/python/BOM_formatting'
wb1 = xl.load_workbook('Pigment Wts.xlsx')
write = wb1.active
for file in os.listdir(mydir):
    if file.endswith('2022.xlsx'):
        wb2 = xl.load_workbook(file)
        for sheet in wb2.sheetnames:
            if sheet == 'list' or sheet == 'List' or sheet[:3] == 'MOT':
                continue
            for x in range(1,250):
                for y in range(1,30):
                    if str(wb2[sheet].cell(row=x,column=y).value) in pigments:
                        if wb2[sheet].cell(row=x,column=y+3).value == 0 or wb2[sheet].cell(row=x,column=y+3).value == None:
                            print('Empty value found, ignoring data here')
                            continue
                        print('{} {} found in {}'.format(wb2[sheet].cell(row=x,column=y).value,wb2[sheet].cell(row=x,column=y+1).value,sheet.split()[0]))
                        write.cell(row=write.max_row+1,column=1).value = str(wb2[sheet].cell(row=x,column=y).value)
                        write.cell(row=write.max_row,column=2).value = str(wb2[sheet].cell(row=x,column=y+1).value)
                        try:
                            write.cell(row=write.max_row,column=3).value = float(wb2[sheet].cell(row=x,column=y+3).value)
                        except:
                            write.cell(row=write.max_row,column=3).value = wb2[sheet].cell(row=x,column=y+3).value
                        write.cell(row=write.max_row,column=4).value = str(file)
                        write.cell(row=write.max_row,column=5).value = str(sheet.split()[0])
                    # if str(wb2[sheet].cell(row=x,column=y).value) in whitepigids:
                    #     wb2[sheet].cell(row=x,column=y).value = '145834'
                        
                    # if str(wb2[sheet].cell(row=x,column=y).value) in whitepignames:
                    #     wb2[sheet].cell(row=x,column=y).value = 'White Pigment Tronox CR834'
    # if file.endswith('2022.xlsx'):
    #     wb2.save(file)
wb1.save('Commercial Pigment Data.xlsx')
# wb = xl.load_workbook('L1 BOMs 2022.xlsx')
# check = True
# for i in wb.sheetnames:
#     print('')
#     print(i, end=' ')
#     if check:
#         check = False
#         continue
#     for row in [4]:
#         print(wb[i].cell(row = row, column = 4).value, end=' ')
