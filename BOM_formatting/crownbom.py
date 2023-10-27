import openpyxl as xl

wb = xl.load_workbook('L5 Crown BOMs 2022.xlsx')
for i in wb.sheetnames:
    print(i, wb[i].cell( row = 18, column = 13).value)
