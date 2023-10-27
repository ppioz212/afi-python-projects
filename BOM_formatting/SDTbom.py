import openpyxl as xl

wb = xl.load_workbook('L5 SDT BOMs 2022.xlsx')
r = 2
c = 1
for i in wb.sheetnames:
    if c == 1 or c == 2:
        c += 1
        continue
    wb['List'].cell(row=r, column=1).value = i
    print(i, wb[i].cell(row=16, column=12).value)
    r += 1
wb.save('L5 SDT BOMs 2022.xlsx')
