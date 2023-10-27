import openpyxl as xl
import os

wb1 = xl.load_workbook('scraps.xlsx')
read = wb1['L5']
mydir = 'C:/Users/asuresh/python'
for file in os.listdir(mydir):
    if file.endswith('7-1-22.xlsx'):
        wb2 = xl.load_workbook(file)
        if file.endswith('L5 Crown BOMs 7-1-22.xlsx') or file.endswith('L5 Chromaspin BOMs 7-1-22.xlsx'):
            print("chromaspin or crown found")
            s_row = 7
        else:
            print("standard found")
            s_row = 6
        for r in range(1,125):
            for sheet in wb2.sheetnames:
                if read.cell(row=r,column=5).value == "Yes":
                    if read.cell(row=r,column=1).value == sheet.split()[0]:
                        # print("Pattern " + sheet.split()[0] + " found")
                        wb2[sheet].cell(row=s_row,column=18).value = round((read.cell(row=r,column=2).value)/10)*10
        wb2.save(file)
