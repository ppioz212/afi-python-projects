import openpyxl as xl
wb = xl.load_workbook('LIMESTONE - SMI.xlsx')
sheet = wb['SMI DATA']   
col = 4
for i in range(1,7000):
	sheet.cell(row=i,column=col).value
	try:
		if sheet.cell(row=i,column=col).value[:3] == 'S0U':
			sheet.cell(row=i,column=col).value = sheet.cell(row=i,column=col).value[0] + 'O' + sheet.cell(row=i,column=col).value[2:]
			print('Change made on Row{}'.format(i))
		elif sheet.cell(row=i,column=col).value[:5] == 'GACxK' or sheet.cell(row=i,column=col).value[:5] == 'GACxX':
			sheet.cell(row=i,column=col).value = sheet.cell(row=i,column=col).value[:3] + 'X' + sheet.cell(row=i,column=col).value[5:]
			print('Change made on Row{}'.format(i))
		elif sheet.cell(row=i,column=col).value[:4] == 'GACK':
			sheet.cell(row=i,column=col).value = sheet.cell(row=i,column=col).value[:3] + 'X' + sheet.cell(row=i,column=col).value[4:]
			print('Change made on Row{}'.format(i))
		elif sheet.cell(row=i,column=col).value[:4] == 'GAXC':
			sheet.cell(row=i,column=col).value = sheet.cell(row=i,column=col).value[:2] + 'CX' + sheet.cell(row=i,column=col).value[4:]
			print('Change made on Row{}'.format(i))
	except:
		pass

wb.save('LIMESTONE - SMI.xlsx')